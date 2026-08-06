import sys
import openpyxl
import urllib.request
import urllib.error
import json
import time

sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = r'C:\Users\HP\.gemini\antigravity\scratch\lanna_dict_extractor\lanna_dictionary_full.xlsx'
API_URL = 'https://siripaporn.lnw.mn/endpoints/vocabulary_api.php?action=create'

# First delete test entry if exists
def delete_test():
    try:
        req = urllib.request.Request(
            'https://siripaporn.lnw.mn/endpoints/vocabulary_api.php?action=delete&id=V00001',
            data=json.dumps({}).encode('utf-8'),
            headers={'Content-Type': 'application/json'}
        )
        res = urllib.request.urlopen(req)
        print('Deleted test entry:', res.read().decode('utf-8'))
    except Exception as e:
        print('Note:', e)

def import_dictionary():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    
    total = ws.max_row - 1  # exclude header
    print(f'Total entries to import: {total}')
    
    success = 0
    failed = 0
    skipped = 0
    
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        page_num, thai_word, reading, meaning = row
        
        # Skip empty rows
        if not thai_word or str(thai_word).strip() == '':
            skipped += 1
            continue
        
        thai_word = str(thai_word).strip()
        reading = str(reading).strip() if reading else ''
        meaning = str(meaning).strip() if meaning else ''
        
        # Limit meaning length to prevent DB issues
        if len(meaning) > 1000:
            meaning = meaning[:1000] + '...'
        
        body = {
            'thai_word': thai_word,
            'lanna_word': '',  # No lanna unicode in xlsx - AI will fill later
            'reading': reading,
            'meaning': meaning,
            'category_vocab_id': None  # No category in xlsx
        }
        
        try:
            req = urllib.request.Request(
                API_URL,
                data=json.dumps(body).encode('utf-8'),
                headers={'Content-Type': 'application/json'}
            )
            res = urllib.request.urlopen(req, timeout=10)
            result = json.loads(res.read().decode('utf-8'))
            success += 1
            if success % 100 == 0:
                print(f'Progress: {success}/{total} imported ({failed} failed, {skipped} skipped)')
        except urllib.error.HTTPError as e:
            err = e.read().decode('utf-8')
            failed += 1
            if failed <= 5:
                print(f'Error on "{thai_word}": {err}')
        except Exception as e:
            failed += 1
            if failed <= 5:
                print(f'Exception on "{thai_word}": {e}')
        
        # Small delay to avoid overwhelming the server
        time.sleep(0.05)
    
    print(f'\nDone! Success: {success}, Failed: {failed}, Skipped: {skipped}')

if __name__ == '__main__':
    delete_test()
    import_dictionary()
